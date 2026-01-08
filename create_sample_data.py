import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Patient Appointments"

# Define styles
header_fill = PatternFill(start_color="0052CC", end_color="0052CC", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sample data
conditions = ["HIV", "Diabetes", "Hypertension", "Maternal Health", "Asthma", "Mixed (HIV+Diabetes)"]
statuses = ["Scheduled", "Confirmed", "No-Show", "Completed", "Rescheduled", "Cancelled"]
regions = ["Kampala", "Fort Portal", "Mbarara", "Jinja", "Arua", "Masaka", "Soroti"]

# Headers
headers = [
    "Patient ID",
    "Full Name",
    "Age",
    "Gender",
    "Phone",
    "Email",
    "Primary Condition",
    "Last Appointment",
    "Next Appointment",
    "Appointment Time",
    "Status",
    "Notes",
    "Region",
    "ART Adherence %",
    "Viral Load Status"
]

# Add headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Generate sample data
base_date = datetime.now()
first_names = ["James", "Mary", "John", "Grace", "Peter", "Anna", "David", "Sarah", "Michael", "Rebecca",
               "Henry", "Linda", "Charles", "Alice", "Robert", "Jane", "Paul", "Elizabeth", "Joseph", "Margaret"]
last_names = ["Okoro", "Mugisha", "Okonkwo", "Kamau", "Adeyemi", "Nkosi", "Mwangi", "Mensah", "Kaunda", "Banda",
              "Kibaki", "Nyerere", "Mandela", "Dlamini", "Achieng", "Obi", "Senghor", "Kenyatta", "Habyarimana", "Chiluba"]

for row in range(2, 52):  # 50 patient records
    col = 1
    
    # Patient ID
    ws.cell(row=row, column=col).value = f"PAT{1000+row}"
    col += 1
    
    # Full Name
    full_name = f"{random.choice(first_names)} {random.choice(last_names)}"
    ws.cell(row=row, column=col).value = full_name
    col += 1
    
    # Age
    age = random.randint(18, 75)
    ws.cell(row=row, column=col).value = age
    col += 1
    
    # Gender
    gender = random.choice(["M", "F"])
    ws.cell(row=row, column=col).value = gender
    col += 1
    
    # Phone
    phone = f"+256{random.randint(7, 9)}{random.randint(10000000, 99999999)}"
    ws.cell(row=row, column=col).value = phone
    col += 1
    
    # Email
    email = f"{full_name.lower().replace(' ', '.')}{random.randint(1, 999)}@email.com"
    ws.cell(row=row, column=col).value = email
    col += 1
    
    # Primary Condition
    condition = random.choice(conditions)
    ws.cell(row=row, column=col).value = condition
    col += 1
    
    # Last Appointment (past dates)
    last_apt = base_date - timedelta(days=random.randint(5, 90))
    ws.cell(row=row, column=col).value = last_apt.strftime("%Y-%m-%d")
    col += 1
    
    # Next Appointment (future dates)
    next_apt = base_date + timedelta(days=random.randint(1, 60))
    ws.cell(row=row, column=col).value = next_apt.strftime("%Y-%m-%d")
    col += 1
    
    # Appointment Time
    time = random.choice(["09:00 AM", "10:30 AM", "02:00 PM", "03:30 PM", "11:00 AM"])
    ws.cell(row=row, column=col).value = time
    col += 1
    
    # Status
    status = random.choice(statuses)
    ws.cell(row=row, column=col).value = status
    col += 1
    
    # Notes
    notes_options = [
        "Stable condition",
        "Needs follow-up lab tests",
        "Medication adjustment required",
        "Patient reports improvement",
        "Requires counseling",
        "Viral suppression achieved",
        "BP monitoring needed"
    ]
    ws.cell(row=row, column=col).value = random.choice(notes_options)
    col += 1
    
    # Region
    region = random.choice(regions)
    ws.cell(row=row, column=col).value = region
    col += 1
    
    # ART Adherence %
    if "HIV" in condition:
        adherence = random.randint(60, 100)
        ws.cell(row=row, column=col).value = f"{adherence}%"
    else:
        ws.cell(row=row, column=col).value = "N/A"
    col += 1
    
    # Viral Load Status
    if "HIV" in condition:
        viral = random.choice(["Undetectable", "Low", "Moderate", "High"])
        ws.cell(row=row, column=col).value = viral
    else:
        ws.cell(row=row, column=col).value = "N/A"
    
    # Apply borders to all cells
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = border
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="left", vertical="center")

# Adjust column widths
column_widths = [12, 20, 8, 8, 15, 25, 20, 15, 15, 15, 15, 20, 12, 15, 18]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Save file
file_path = "c:/Users/kitwe/Desktop/IHM/Patient_Appointments_Sample_Data.xlsx"
wb.save(file_path)
print(f"✓ Excel file created successfully: {file_path}")
print(f"✓ Total records: 50 patients")
print(f"✓ Columns: {len(headers)}")
